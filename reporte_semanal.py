"""
reporte_semanal.py (v11)
=======================
Genera un Excel SEMANAL con hojas separadas:
- Semana (Lun-Vie) - solo programas con datos
- Fin de Semana (Sáb-Dom) - solo programas con datos
- YouTube (con desglose por día)
- Datos Crudos

3X3 se trata igual que el resto de programas, salvo que se omite en fin de semana (sáb/dom).
"""

import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from googleapiclient.discovery import build
from datetime import datetime, timedelta, timezone
import re, os, glob

from config_local import CARPETA_DATOS, YOUTUBE_API_KEY, PROGRAMAS_CONFIG, MYSQL_CONFIG, MYSQL_TABLA

def detectar_semana():
    archivos = {}
    for fp in sorted(glob.glob(os.path.join(CARPETA_DATOS, "*_programacion.xlsx"))):
        m = re.match(r".*?(\d{4}-\d{2}-\d{2})_programacion\.xlsx", fp)
        if m:
            archivos[m.group(1)] = fp
    if not archivos:
        return None, []

    fechas_set = set(archivos.keys())
    fechas_ord = sorted(fechas_set)

    # Buscar la semana Lun–Dom más reciente que tenga los 7 días disponibles
    for fecha_str in reversed(fechas_ord):
        dt = datetime.strptime(fecha_str, "%Y-%m-%d")
        # Retroceder hasta el lunes de esa semana
        lunes = dt - timedelta(days=dt.weekday())
        semana = [(lunes + timedelta(days=i)).strftime("%Y-%m-%d") for i in range(7)]
        if all(f in fechas_set for f in semana):
            return semana, [archivos[f] for f in semana]

    # Fallback: cualquier racha de 7 días consecutivos más reciente
    if len(fechas_ord) >= 7:
        for i in range(len(fechas_ord) - 6, -1, -1):
            candidates = fechas_ord[i:i+7]
            dates_dt = [datetime.strptime(f, "%Y-%m-%d") for f in candidates]
            gaps = [(dates_dt[j+1] - dates_dt[j]).days for j in range(6)]
            if all(g == 1 for g in gaps):
                return candidates, [archivos[f] for f in candidates]

    return fechas_ord, [archivos[f] for f in fechas_ord]

def detectar_programas_s3(filepaths):
    programas = {}
    for fp in filepaths:
        try:
            df = pd.read_excel(fp)
            col_titulo = None
            for col in df.columns:
                if col.lower().strip() in ["título", "titulo"]:
                    col_titulo = col
                    break
            if not col_titulo:
                continue
            for _, row in df.iterrows():
                titulo = str(row.get(col_titulo, "")).strip()
                if titulo.upper() != "TOTAL" and titulo and str(titulo).lower() != "nan":
                    programas[titulo] = True
        except Exception as e:
            print(f"  ⚠️  {fp}: {e}")
    return list(programas.keys())

def normalizar_programa(titulo):
    """Agrupa variantes de programas"""
    titulo = str(titulo).strip()
    
    # Agrupar Vecinos al Límite:
    # - (especial)  → Vecinos al Límite
    # - (lo mejor)  → Vecinos al Límite Extra
    # - Vecinos al Límite / (resumen) / Extra → sin cambio (son programas separados)
    if "VECINOS" in titulo.upper() and "LIMITE" in titulo.upper():
        if "ESPECIAL" in titulo.upper():
            return "Vecinos al Límite"
        elif "LO MEJOR" in titulo.upper():
            return "Vecinos al Límite Extra"
    
    return titulo

def obtener_datos_s3(fechas, filepaths, programas):
    # Normalizar nombres de programas
    programas_normalizados = {}
    for p in programas:
        norm = normalizar_programa(p)
        if norm not in programas_normalizados:
            programas_normalizados[norm] = []
        programas_normalizados[norm].append(p)

    datos = {prog: {} for prog in programas_normalizados.keys()}
    datos_crudos = []

    for fecha, fp in zip(fechas, filepaths):
        dt_fecha = datetime.strptime(fecha, "%Y-%m-%d")
        es_finde = dt_fecha.weekday() >= 5
        df = pd.read_excel(fp)

        col_titulo = None
        col_plays = None
        for col in df.columns:
            col_lower = col.lower().strip()
            if col_lower in ["título", "titulo"]:
                col_titulo = col
            elif col_lower == "plays":
                col_plays = col

        if not col_titulo or not col_plays:
            continue

        for prog_norm, prog_originals in programas_normalizados.items():
            for prog_original in prog_originals:
                # 3X3 no se considera en fin de semana
                if "3 X 3" in prog_original.upper() and es_finde:
                    continue

                plays_list = []
                for _, row in df.iterrows():
                    titulo = str(row.get(col_titulo, "")).strip()
                    if titulo == prog_original:
                        plays_val = row.get(col_plays, 0)
                        try:
                            if pd.isna(plays_val) or plays_val == "" or str(plays_val).lower() == "nan":
                                plays_val = 0
                            else:
                                plays_val = int(float(plays_val))
                            plays_list.append(plays_val)
                        except:
                            plays_list.append(0)

                if plays_list:
                    valor = sum(plays_list)
                    if prog_norm not in datos:
                        datos[prog_norm] = {}
                    datos[prog_norm][fecha] = valor
                    datos_crudos.append({"fecha": fecha, "programa": prog_norm, "plays": valor, "fuente": "S3"})
                    break

    return datos, datos_crudos

TZ_CHILE = timezone(timedelta(hours=-4))

def utc_a_fecha_chile(pub_utc):
    """Convierte publishedAt de YouTube (UTC) a fecha local Chile (UTC-4)"""
    if not pub_utc:
        return ""
    try:
        dt = datetime.fromisoformat(pub_utc.replace("Z", "+00:00"))
        return dt.astimezone(TZ_CHILE).strftime("%Y-%m-%d")
    except Exception:
        return pub_utc[:10]

def obtener_views_playlist_por_dia(yt, playlist_id, fecha_ini, fecha_fin):
    """Obtiene views por video usando la fecha real de publicación del video (no la de la playlist)"""
    # Paso 1: recolectar todos los video_ids de la playlist
    todos_video_ids = []
    page_token = None
    try:
        while True:
            resp = yt.playlistItems().list(part="snippet", playlistId=playlist_id, maxResults=50, pageToken=page_token).execute()
            for item in resp.get("items", []):
                video_id = item.get("snippet", {}).get("resourceId", {}).get("videoId")
                if video_id:
                    todos_video_ids.append(video_id)
            page_token = resp.get("nextPageToken")
            if not page_token:
                break

        # Paso 2: obtener snippet+statistics de cada video para usar fecha real de publicación
        videos_por_fecha = {}
        for i in range(0, len(todos_video_ids), 50):
            chunk = todos_video_ids[i:i+50]
            resp = yt.videos().list(part="snippet,statistics", id=",".join(chunk)).execute()
            for item in resp.get("items", []):
                pub_utc = item.get("snippet", {}).get("publishedAt", "")
                publicado = utc_a_fecha_chile(pub_utc)
                if fecha_ini <= publicado <= fecha_fin:
                    views = int(item["statistics"].get("viewCount", 0))
                    videos_por_fecha[publicado] = videos_por_fecha.get(publicado, 0) + views

        return videos_por_fecha
    except Exception as e:
        print(f"    ⚠️  Playlist: {e}")
        return {}

def obtener_datos_youtube(fechas):
    resultado = {}
    datos_crudos_yt = []
    if not YOUTUBE_API_KEY or YOUTUBE_API_KEY == "TU_API_KEY_AQUI":
        print("   ⚠️  Sin API key")
        return resultado, datos_crudos_yt
    try:
        yt = build("youtube", "v3", developerKey=YOUTUBE_API_KEY)
        fecha_ini = min(fechas)
        fecha_fin = max(fechas)
        for prog, cfg in PROGRAMAS_CONFIG.items():
            playlist_raw = cfg.get("playlist_youtube")
            if not playlist_raw:
                continue
            playlists = playlist_raw if isinstance(playlist_raw, list) else [playlist_raw]
            print(f"   🎥 {prog}...", end="")
            views_por_dia = {}
            for pl_id in playlists:
                for fecha, views in obtener_views_playlist_por_dia(yt, pl_id, fecha_ini, fecha_fin).items():
                    views_por_dia[fecha] = views_por_dia.get(fecha, 0) + views
            if views_por_dia:
                views_total = sum(views_por_dia.values())
                resultado[prog] = views_por_dia
                datos_crudos_yt.append({"fecha": "N/A", "programa": prog, "plays": views_total, "fuente": "YouTube"})
                print(f" ✅ {views_total:,.0f}")
            else:
                print(" (sin datos)")
    except Exception as e:
        print(f"   ❌ {e}")
    return resultado, datos_crudos_yt

def label_dia(fecha_str):
    dt = datetime.strptime(fecha_str, "%Y-%m-%d")
    dias = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]
    return f"{dias[dt.weekday()]} {dt.strftime('%d/%m')}"

def crear_hoja_s3(wb, fechas, s3_datos, titulo, fechas_filter):
    """Crea una hoja de S3 (Semana o Fin de Semana) - solo con programas que tienen datos"""
    AZUL_OSC, AZUL_MED, VERDE, VERDE_BG, AMARILLO, BLANCO, GRIS_CLAR = "1F3864", "2E5FA3", "375623", "E2EFDA", "FFF2CC", "FFFFFF", "F5F5F5"
    
    ws = wb.create_sheet(titulo)
    rango = f"{label_dia(fechas[0])} al {label_dia(fechas[-1])}"
    
    ws.merge_cells("A1:I1")
    ws["A1"] = titulo.upper()
    ws["A1"].font = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", start_color=AZUL_OSC, fgColor=AZUL_OSC)
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:I2")
    ws["A2"] = f"Período: {rango}"
    ws["A2"].font = Font(name="Calibri", size=10, color="666666", italic=True)
    ws["A2"].fill = PatternFill("solid", start_color="EEF2F8", fgColor="EEF2F8")
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 8

    # Headers
    headers = ["Programa", "Total", "Promedio/día", "Máximo", "Mínimo"] + [label_dia(f) for f in fechas_filter]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=ci, value=h)
        c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill("solid", start_color=AZUL_MED, fgColor=AZUL_MED)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[4].height = 32
    
    ws.column_dimensions["A"].width = 22
    for i in range(1, 5):
        ws.column_dimensions[chr(65+i)].width = 13
    for i in range(len(fechas_filter)):
        ws.column_dimensions[chr(70 + i)].width = 12

    # Datos - SOLO programas con datos en este período
    programas_con_datos = []
    for prog in s3_datos:
        tiene_datos_en_periodo = any(s3_datos[prog].get(f, 0) > 0 for f in fechas_filter)
        if tiene_datos_en_periodo:
            programas_con_datos.append((prog, s3_datos[prog]))
    
    # Ordenar por total
    programas_con_datos = sorted(programas_con_datos, 
                                 key=lambda x: sum(x[1].get(f, 0) for f in fechas_filter), reverse=True)
    
    for pi, (prog, dias_data) in enumerate(programas_con_datos):
        fila = 5 + pi
        vals = [dias_data.get(f, 0) for f in fechas_filter]
        activos = [v for v in vals if v > 0]
        total = sum(activos)
        promedio = int(total / len(activos)) if activos else 0
        maximo = max(activos) if activos else 0
        minimo = min(activos) if activos else 0
        bg = BLANCO if pi % 2 == 0 else GRIS_CLAR

        ws.cell(row=fila, column=1, value=prog).fill = PatternFill("solid", start_color=bg, fgColor=bg)
        for col, val in enumerate([total, promedio, maximo, minimo], 2):
            c = ws.cell(row=fila, column=col, value=val if val else None)
            c.number_format = '#,##0'
            c.font = Font(name="Calibri", size=10, bold=(col == 2), color=VERDE if col == 2 else "000000")
            c.fill = PatternFill("solid", start_color=VERDE_BG if col == 2 else bg, fgColor=VERDE_BG if col == 2 else bg)
            c.alignment = Alignment(horizontal="right")
        for di, val in enumerate(vals):
            c = ws.cell(row=fila, column=6 + di, value=val if val else None)
            c.number_format = '#,##0'
            bg_dia = AMARILLO if val > 0 else bg
            c.fill = PatternFill("solid", start_color=bg_dia, fgColor=bg_dia)
            c.alignment = Alignment(horizontal="right")

    # Total por día
    fila_tot = 5 + len(programas_con_datos)
    c = ws.cell(row=fila_tot, column=1, value="TOTAL")
    c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    c.fill = PatternFill("solid", start_color=AZUL_OSC, fgColor=AZUL_OSC)
    for col in range(2, 6):
        ws.cell(row=fila_tot, column=col).fill = PatternFill("solid", start_color=AZUL_OSC, fgColor=AZUL_OSC)
    for di in range(len(fechas_filter)):
        total_dia = sum(s3_datos[p].get(fechas_filter[di], 0) for p, _ in programas_con_datos)
        c = ws.cell(row=fila_tot, column=6 + di, value=total_dia)
        c.number_format = '#,##0'
        c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill("solid", start_color=AZUL_OSC, fgColor=AZUL_OSC)
        c.alignment = Alignment(horizontal="right")

def crear_excel(fechas, s3_datos, yt_datos, datos_crudos):
    AZUL_OSC, AZUL_MED, VERDE, VERDE_BG, BLANCO, GRIS_CLAR = "1F3864", "2E5FA3", "375623", "E2EFDA", "FFFFFF", "F5F5F5"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Separar semana y fin de semana
    fechas_semana = []
    fechas_fin = []
    
    for f in fechas:
        dt = datetime.strptime(f, "%Y-%m-%d")
        is_finde = dt.weekday() >= 5
        if is_finde:
            fechas_fin.append(f)
        else:
            fechas_semana.append(f)

    # HOJA 1: Semana (Lun-Vie)
    if fechas_semana:
        crear_hoja_s3(wb, fechas, s3_datos, "📅 Semana (Lun-Vie)", fechas_semana)

    # HOJA 2: Fin de Semana (Sáb-Dom)
    if fechas_fin:
        crear_hoja_s3(wb, fechas, s3_datos, "🏖️ Fin de Semana (Sáb-Dom)", fechas_fin)

    # HOJA 3: YouTube
    ws2 = wb.create_sheet("📺 YouTube")
    rango = f"{label_dia(fechas[0])} al {label_dia(fechas[-1])}"
    
    ws2.merge_cells("A1:I1")
    ws2["A1"] = "RESUMEN YOUTUBE — VIEWS POR PROGRAMA Y DÍA"
    ws2["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws2["A1"].fill = PatternFill("solid", start_color=AZUL_OSC, fgColor=AZUL_OSC)
    ws2.row_dimensions[1].height = 28

    ws2.merge_cells("A2:I2")
    ws2["A2"] = f"Período: {rango}"
    ws2["A2"].font = Font(name="Calibri", size=9, color="666666", italic=True)
    ws2["A2"].fill = PatternFill("solid", start_color="EEF2F8", fgColor="EEF2F8")
    ws2.row_dimensions[2].height = 18
    ws2.row_dimensions[3].height = 8

    headers_yt = ["Programa", "Total Semana"] + [label_dia(f) for f in fechas]
    for ci, h in enumerate(headers_yt, 1):
        c = ws2.cell(row=4, column=ci, value=h)
        c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill("solid", start_color=AZUL_MED, fgColor=AZUL_MED)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws2.row_dimensions[4].height = 32

    yt_programas = sorted([(p, yt_datos[p]) for p in yt_datos], 
                          key=lambda x: sum(x[1].values()) or 0, reverse=True)
    
    for ri, (prog, views_por_dia) in enumerate(yt_programas, 1):
        fila = 4 + ri
        bg = BLANCO if ri % 2 == 1 else GRIS_CLAR
        
        c1 = ws2.cell(row=fila, column=1, value=prog)
        c1.fill = PatternFill("solid", start_color=bg, fgColor=bg)
        c1.font = Font(name="Calibri", size=10)
        
        total_views = sum(views_por_dia.values())
        c_tot = ws2.cell(row=fila, column=2, value=total_views)
        c_tot.number_format = '#,##0'
        c_tot.fill = PatternFill("solid", start_color=VERDE_BG, fgColor=VERDE_BG)
        c_tot.font = Font(name="Calibri", bold=True, size=10, color=VERDE)
        c_tot.alignment = Alignment(horizontal="right")
        
        for di, fecha in enumerate(fechas):
            views_dia = views_por_dia.get(fecha, 0)
            c = ws2.cell(row=fila, column=3 + di, value=views_dia if views_dia else None)
            c.number_format = '#,##0'
            c.fill = PatternFill("solid", start_color=bg, fgColor=bg)
            c.alignment = Alignment(horizontal="right")

    fila_tot = 4 + len(yt_programas) + 1
    c_tot_label = ws2.cell(row=fila_tot, column=1, value="TOTAL DIARIO")
    c_tot_label.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    c_tot_label.fill = PatternFill("solid", start_color=AZUL_OSC, fgColor=AZUL_OSC)
    ws2.cell(row=fila_tot, column=2).fill = PatternFill("solid", start_color=AZUL_OSC, fgColor=AZUL_OSC)
    
    for di, fecha in enumerate(fechas):
        total_dia = sum(yt_datos[p].get(fecha, 0) for p in yt_datos)
        c = ws2.cell(row=fila_tot, column=3 + di, value=total_dia)
        c.number_format = '#,##0'
        c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill("solid", start_color=AZUL_OSC, fgColor=AZUL_OSC)
        c.alignment = Alignment(horizontal="right")

    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 14
    for i in range(len(fechas)):
        ws2.column_dimensions[chr(67 + i)].width = 12

    # HOJA 4: Datos Crudos
    ws3 = wb.create_sheet("📋 Datos Crudos")
    ws3.merge_cells("A1:E1")
    ws3["A1"] = "LISTADO COMPLETO DE DATOS"
    ws3["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws3["A1"].fill = PatternFill("solid", start_color=AZUL_OSC, fgColor=AZUL_OSC)
    ws3.row_dimensions[1].height = 28

    headers_crudos = ["Fecha", "Día", "Programa", "Cifra", "Fuente"]
    for ci, h in enumerate(headers_crudos, 1):
        c = ws3.cell(row=2, column=ci, value=h)
        c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill("solid", start_color=AZUL_MED, fgColor=AZUL_MED)
        c.alignment = Alignment(horizontal="center")
    ws3.row_dimensions[2].height = 24

    dias_es = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
    fila_datos = 3
    for registro in sorted(datos_crudos, key=lambda x: (x["fecha"], x["programa"])):
        bg = BLANCO if fila_datos % 2 == 0 else GRIS_CLAR
        ws3.cell(row=fila_datos, column=1, value=registro["fecha"]).fill = PatternFill("solid", start_color=bg, fgColor=bg)
        if registro["fecha"] != "N/A":
            dt = datetime.strptime(registro["fecha"], "%Y-%m-%d")
            ws3.cell(row=fila_datos, column=2, value=dias_es[dt.weekday()]).fill = PatternFill("solid", start_color=bg, fgColor=bg)
        else:
            ws3.cell(row=fila_datos, column=2, value="N/A").fill = PatternFill("solid", start_color=bg, fgColor=bg)
        
        ws3.cell(row=fila_datos, column=3, value=registro["programa"]).fill = PatternFill("solid", start_color=bg, fgColor=bg)
        c = ws3.cell(row=fila_datos, column=4, value=registro["plays"])
        c.number_format = '#,##0'
        c.fill = PatternFill("solid", start_color=bg, fgColor=bg)
        c.alignment = Alignment(horizontal="right")
        ws3.cell(row=fila_datos, column=5, value=registro["fuente"]).fill = PatternFill("solid", start_color=bg, fgColor=bg)
        fila_datos += 1

    ws3.column_dimensions["A"].width = 13
    ws3.column_dimensions["B"].width = 13
    ws3.column_dimensions["C"].width = 26
    ws3.column_dimensions["D"].width = 14
    ws3.column_dimensions["E"].width = 10
    ws3.auto_filter.ref = f"A2:E{fila_datos - 1}"

    return wb

def preparar_registros_bd(datos_crudos_s3, yt_datos):
    """Genera registros por día para S3 y YouTube (YouTube con desglose diario)"""
    registros = list(datos_crudos_s3)
    for prog, views_por_dia in yt_datos.items():
        for fecha, views in views_por_dia.items():
            if views > 0:
                registros.append({"fecha": fecha, "programa": prog, "plays": views, "fuente": "YouTube"})
    return registros

def guardar_en_mysql(registros, fechas):
    try:
        import mysql.connector
    except ImportError:
        print("   ❌ mysql-connector-python no instalado. Ejecuta: pip install mysql-connector-python")
        return False

    try:
        cx = mysql.connector.connect(**MYSQL_CONFIG)
        cur = cx.cursor()

        cur.execute(f"""
            CREATE TABLE IF NOT EXISTS `{MYSQL_TABLA}` (
                id          INT AUTO_INCREMENT PRIMARY KEY,
                fecha       DATE        NOT NULL,
                programa    VARCHAR(120) NOT NULL,
                plays       INT         NOT NULL DEFAULT 0,
                fuente      VARCHAR(20) NOT NULL,
                procesado   DATETIME    DEFAULT CURRENT_TIMESTAMP,
                UNIQUE KEY uk_fecha_prog_fuente (fecha, programa, fuente)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """)

        # Borrar registros existentes de la semana para permitir reprocesamiento
        placeholders = ",".join(["%s"] * len(fechas))
        cur.execute(f"DELETE FROM `{MYSQL_TABLA}` WHERE fecha IN ({placeholders})", fechas)
        eliminados = cur.rowcount

        # Insertar registros nuevos
        sql = f"""
            INSERT INTO `{MYSQL_TABLA}` (fecha, programa, plays, fuente)
            VALUES (%s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE plays = VALUES(plays), procesado = CURRENT_TIMESTAMP
        """
        filas = [(r["fecha"], r["programa"], r["plays"], r["fuente"])
                 for r in registros if r["fecha"] != "N/A"]
        cur.executemany(sql, filas)
        cx.commit()
        print(f"   ✅ {cur.rowcount} registros grabados ({eliminados} anteriores reemplazados)")
        cur.close()
        cx.close()
        return True

    except Exception as e:
        print(f"   ❌ Error MySQL: {e}")
        return False

def main():
    try:
        print(f"\n📂 Carpeta: {CARPETA_DATOS}")
        fechas, filepaths = detectar_semana()
        if not fechas:
            print("❌ No hay archivos *_programacion.xlsx")
            return

        print(f"✅ Semana detectada: {fechas[0]} → {fechas[-1]} ({len(fechas)} días)")
        
        print("\n🔍 Detectando programas en S3...")
        programas = detectar_programas_s3(filepaths)
        print(f"   ✅ {len(programas)} programas encontrados")
        if not programas:
            print("   ⚠️  Sin programas")
            return

        print("\n📊 Extrayendo datos S3...")
        s3_datos, datos_crudos_s3 = obtener_datos_s3(fechas, filepaths, programas)
        print(f"   ✅ {len([p for p in s3_datos if s3_datos[p]])} con datos")

        print("\n📺 Consultando YouTube...")
        yt_datos, datos_crudos_yt = obtener_datos_youtube(fechas)
        print(f"   ✅ {len(yt_datos)} programas con datos")

        datos_crudos = datos_crudos_s3 + datos_crudos_yt

        print("\n📝 Generando Excel...")
        wb = crear_excel(fechas, s3_datos, yt_datos, datos_crudos)

        dt_ini = datetime.strptime(fechas[0], "%Y-%m-%d")
        dt_fin = datetime.strptime(fechas[-1], "%Y-%m-%d")
        nombre = f"Reporte_{dt_ini.strftime('%d-%b')}_{dt_fin.strftime('%d-%b_%Y')}.xlsx"
        salida = os.path.join(CARPETA_DATOS, nombre)

        wb.save(salida)
        print(f"\n✅ Reporte listo:")
        print(f"   📄 {nombre}")
        print(f"   📍 {CARPETA_DATOS}")

        print("\n💾 ¿Grabar datos en la base de datos MySQL? (s/n): ", end="", flush=True)
        respuesta = input().strip().lower()
        if respuesta in ("s", "si", "sí", "y", "yes"):
            print(f"\n💾 Grabando en MySQL ({MYSQL_CONFIG['database']}.{MYSQL_TABLA})...")
            registros_bd = preparar_registros_bd(datos_crudos_s3, yt_datos)
            guardar_en_mysql(registros_bd, fechas)
        else:
            print("   ⏭️  Grabación omitida.")

    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
