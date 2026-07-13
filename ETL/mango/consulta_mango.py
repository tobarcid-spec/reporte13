"""
consulta_mango.py
=================
Consulta la API Mango para cualquier señal, rango de fechas y horario.
Exporta los resultados a Excel.

Uso:
    python consulta_mango.py
"""

import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

import requests, time, os
from datetime import date, datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ─── Configuración ────────────────────────────────────────────────────────────
API_URL = "https://api-mango.digitalproserver.com/"
TOKEN   = "9feb892a7cc282a6829354b7db9449afeeeb39e8eb5b4c4e2a94532bedc2c487"

SENALES_CONOCIDAS = {
    "1": "T13",
    "2": "C13",
    "3": "CNN_CHILE",
    "4": "T13R",
}

CARPETA_SALIDA = r"C:\procesos\ReporteCanal13\datos"

DIAS_ES = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]

# ─── Helpers ──────────────────────────────────────────────────────────────────

def pedir(msg, default=None):
    sufijo = f" [{default}]" if default else ""
    valor = input(f"{msg}{sufijo}: ").strip()
    return valor if valor else default

def pedir_fecha(msg):
    while True:
        txt = input(f"{msg} (YYYY-MM-DD): ").strip()
        try:
            return datetime.strptime(txt, "%Y-%m-%d").date()
        except ValueError:
            print("  Formato incorrecto. Ej: 2026-06-11")

def pedir_hora(msg):
    while True:
        txt = input(f"{msg} (HH:MM): ").strip()
        try:
            datetime.strptime(txt, "%H:%M")
            return txt + ":00"
        except ValueError:
            print("  Formato incorrecto. Ej: 22:25")

def fechas_en_rango(desde: date, hasta: date):
    d = desde
    while d <= hasta:
        yield d
        d += timedelta(days=1)

def consultar_mango(slug, date_start, date_end, reintentos=3):
    payload = {
        "token":      TOKEN,
        "slug":       slug,
        "type":       "overview",
        "format":     "live",
        "date_start": date_start,
        "date_end":   date_end,
    }
    for intento in range(reintentos):
        try:
            r = requests.post(API_URL, json=payload, timeout=60)
            d = r.json()
            if "plays" in d:
                return d
            print(f"  ⚠️  Respuesta inesperada: {d}")
            return None
        except Exception as e:
            if intento < reintentos - 1:
                print(f"  Reintentando ({intento+2}/{reintentos})...")
                time.sleep(3)
            else:
                print(f"  ❌ Error tras {reintentos} intentos: {e}")
                return None

def avg_a_minutos(avg_str):
    """Convierte '0 m, 0 d, 00:06:13 h' a '6:13 min'"""
    try:
        partes = avg_str.split(",")
        hms = partes[-1].strip().replace(" h", "")
        h, m, s = map(int, hms.split(":"))
        total_min = h * 60 + m
        return f"{total_min}:{s:02d} min"
    except Exception:
        return avg_str

# ─── Generación de Excel ──────────────────────────────────────────────────────

def exportar_excel(slug, registros, carpeta):
    AZUL_OSC = "1F3864"; AZUL_MED = "2E5FA3"
    VERDE_BG = "E2EFDA"; VERDE    = "375623"
    BLANCO   = "FFFFFF"; GRIS     = "F5F5F5"
    AMARILLO = "FFF2CC"

    fecha_ini = registros[0]["fecha"]
    fecha_fin = registros[-1]["fecha"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = slug

    # Título
    ws.merge_cells("A1:I1")
    ws["A1"] = f"SEÑAL {slug} — {fecha_ini} al {fecha_fin}"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=AZUL_OSC)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Encabezados
    headers = ["Fecha", "Día", "Inicio", "Fin", "Duración", "Plays", "Streams", "Devices", "Avg Time"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=AZUL_MED)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 24

    # Datos
    for ri, reg in enumerate(registros):
        fila = 3 + ri
        bg = BLANCO if ri % 2 == 0 else GRIS
        vals = [
            reg["fecha"], reg["dia"], reg["inicio"], reg["fin"],
            reg["duracion"], reg["plays"], reg["streams"], reg["devices"], reg["avg_time"]
        ]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=fila, column=ci, value=val)
            c.font = Font(name="Calibri", size=10)
            c.fill = PatternFill("solid", fgColor=bg)
            if ci in (6, 7, 8):
                c.number_format = "#,##0"
                c.alignment = Alignment(horizontal="right")
                if ci == 6:
                    c.fill = PatternFill("solid", fgColor=VERDE_BG)
                    c.font = Font(name="Calibri", size=10, bold=True, color=VERDE)
                elif ci == 7:
                    c.fill = PatternFill("solid", fgColor=AMARILLO)
            else:
                c.alignment = Alignment(horizontal="center")

    # Fila total
    fila_tot = 3 + len(registros)
    ws.merge_cells(f"A{fila_tot}:E{fila_tot}")
    c = ws.cell(row=fila_tot, column=1, value="TOTAL")
    c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=AZUL_OSC)
    c.alignment = Alignment(horizontal="center")
    for ci, campo in [(6, "plays"), (7, "streams"), (8, "devices")]:
        total = sum(r[campo] for r in registros if r[campo])
        c = ws.cell(row=fila_tot, column=ci, value=total)
        c.number_format = "#,##0"
        c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=AZUL_OSC)
        c.alignment = Alignment(horizontal="right")

    # Anchos
    for col, ancho in zip("ABCDEFGHI", [13, 6, 8, 8, 10, 12, 12, 12, 12]):
        ws.column_dimensions[col].width = ancho

    nombre = f"Mango_{slug}_{fecha_ini}_{fecha_fin}.xlsx"
    salida = os.path.join(carpeta, nombre)
    wb.save(salida)
    return salida

# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    print("\n╔══════════════════════════════════════╗")
    print("║   Consulta Mango — Señal / Horario   ║")
    print("╚══════════════════════════════════════╝\n")

    # 1. Selección de señal
    print("Señales conocidas:")
    for k, v in SENALES_CONOCIDAS.items():
        print(f"  {k}. {v}")
    print("  0. Ingresar slug manualmente")

    opcion = pedir("Selecciona una opción").strip()
    if opcion == "0" or opcion not in SENALES_CONOCIDAS:
        slug = pedir("Slug de la señal").upper()
    else:
        slug = SENALES_CONOCIDAS[opcion]
    print(f"\n  ✅ Señal: {slug}")

    # 2. Rango de fechas
    print()
    desde = pedir_fecha("Fecha DESDE")
    hasta = pedir_fecha("Fecha HASTA ")

    # 3. Horario
    print()
    mismo_horario = pedir("¿Mismo horario para todos los días? (s/n)", "s").lower()

    if mismo_horario in ("s", "si", "sí", "y"):
        hora_ini = pedir_hora("Hora INICIO")
        hora_fin = pedir_hora("Hora FIN   ")
        horarios = {str(f): (hora_ini, hora_fin) for f in fechas_en_rango(desde, hasta)}
    else:
        horarios = {}
        for f in fechas_en_rango(desde, hasta):
            fecha_str = str(f)
            dia = DIAS_ES[f.weekday()]
            print(f"\n  {fecha_str} ({dia})")
            horarios[fecha_str] = (pedir_hora("  Inicio"), pedir_hora("  Fin   "))

    # 4. Consultar API
    print(f"\n🔍 Consultando API Mango ({slug})...\n")
    registros = []

    for f in fechas_en_rango(desde, hasta):
        fecha_str = str(f)
        dia = DIAS_ES[f.weekday()]
        ini, fin = horarios[fecha_str]

        print(f"  {fecha_str} ({dia}) {ini[:5]}–{fin[:5]}...", end=" ", flush=True)
        datos = consultar_mango(slug, f"{fecha_str} {ini}", f"{fecha_str} {fin}")

        if datos:
            t_ini = datetime.strptime(ini, "%H:%M:%S")
            t_fin = datetime.strptime(fin, "%H:%M:%S")
            mins  = int((t_fin - t_ini).total_seconds() // 60)
            registros.append({
                "fecha":    fecha_str,
                "dia":      dia,
                "inicio":   ini[:5],
                "fin":      fin[:5],
                "duracion": f"{mins} min",
                "plays":    int(datos.get("plays", 0)),
                "streams":  int(datos.get("streams", 0)),
                "devices":  int(datos.get("devices", 0)),
                "avg_time": avg_a_minutos(datos.get("average_time", "")),
            })
            print(f"Plays: {int(datos['plays']):,}")
        else:
            print("sin datos")

    if not registros:
        print("\n❌ Sin datos para exportar.")
        return

    # 5. Resumen
    print(f"\n{'─'*45}")
    print(f"  Total plays:   {sum(r['plays'] for r in registros):>10,}")
    print(f"  Total streams: {sum(r['streams'] for r in registros):>10,}")
    print(f"  Total devices: {sum(r['devices'] for r in registros):>10,}")
    print(f"{'─'*45}")

    # 6. Exportar Excel
    print("\n📝 Generando Excel...")
    salida = exportar_excel(slug, registros, CARPETA_SALIDA)
    print(f"  ✅ {os.path.basename(salida)}")
    print(f"  📍 {CARPETA_SALIDA}")

if __name__ == "__main__":
    main()
