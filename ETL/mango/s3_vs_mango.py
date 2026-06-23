"""
s3_vs_mango.py
==============
Combina datos de C13 (programacion diaria) con señal de Mango API.

Flujo:
  1. Lee los archivos *_programacion.xlsx del rango de fechas indicado
  2. Busca el programa solicitado y extrae Inicio, Fin y Plays por día
  3. Consulta Mango API con esas franjas horarias para la señal elegida
  4. Une ambas fuentes y exporta a Excel

Uso:
    python s3_vs_mango.py
"""

import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

import os, requests, time
from datetime import date, datetime, timedelta
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ─── Configuración ────────────────────────────────────────────────────────────
API_URL        = "https://api-mango.digitalproserver.com/"
TOKEN          = "9feb892a7cc282a6829354b7db9449afeeeb39e8eb5b4c4e2a94532bedc2c487"
CARPETA_DATOS  = r"C:\procesos\ReporteCanal13\datos"
CARPETA_SALIDA = r"C:\procesos\ReporteCanal13\datos"

SENALES_CONOCIDAS = {
    "1": "T13",
    "2": "C13",
    "3": "CNN_CHILE",
    "4": "T13R",
}

DIAS_ES = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]

# ─── Helpers ──────────────────────────────────────────────────────────────────

def pedir(msg, default=None):
    sufijo = f" [{default}]" if default else ""
    valor = input(f"{msg}{sufijo}: ").strip()
    return valor if valor else default

def pedir_fecha(msg):
    while True:
        txt = input(f"{msg} (YYYY-MM-DD): ").strip()
        try:
            return datetime.strptime(txt, "%Y-%m-%d").date()
        except ValueError:
            print("  Formato incorrecto. Ej: 2026-06-11")

def fechas_en_rango(desde: date, hasta: date):
    d = desde
    while d <= hasta:
        yield d
        d += timedelta(days=1)

def avg_a_minutos(avg_str):
    try:
        partes = avg_str.split(",")
        hms = partes[-1].strip().replace(" h", "")
        h, m, s = map(int, hms.split(":"))
        return f"{h*60+m}:{s:02d} min"
    except Exception:
        return avg_str

# ─── Paso 1: buscar programa en archivos C13 ─────────────────────────────────

def listar_programas_disponibles(desde: date, hasta: date):
    """Retorna set de títulos únicos encontrados en el rango de fechas."""
    programas = set()
    for f in fechas_en_rango(desde, hasta):
        fp = os.path.join(CARPETA_DATOS, f"{f}_programacion.xlsx")
        if not os.path.exists(fp):
            continue
        try:
            df = pd.read_excel(fp)
            col = next((c for c in df.columns if c.lower().strip() in ["título", "titulo"]), None)
            if col:
                for t in df[col].dropna().astype(str):
                    t = t.strip()
                    if t and t.upper() != "TOTAL":
                        programas.add(t)
        except Exception:
            pass
    return sorted(programas)

def buscar_programa_en_s3(desde: date, hasta: date, termino: str):
    """
    Busca el termino en los archivos C13 del rango.
    Retorna lista de dicts: fecha_archivo, fecha_emision, dia, ini_dt, fin_dt, ini, fin, plays_s3.

    IMPORTANTE: los programas nocturnos (después de medianoche) tienen en la columna
    Inicial/Final la fecha REAL de emisión, que puede ser distinta a la del archivo.
    Se usa esa fecha real para consultar Mango correctamente.
    """
    resultados = []
    termino_upper = termino.upper()

    for f in fechas_en_rango(desde, hasta):
        fp = os.path.join(CARPETA_DATOS, f"{f}_programacion.xlsx")
        if not os.path.exists(fp):
            print(f"  ⚠️  Sin archivo para {f}")
            continue
        try:
            df = pd.read_excel(fp)
            col_titulo = next((c for c in df.columns if c.lower().strip() in ["título", "titulo"]), None)
            col_plays  = next((c for c in df.columns if c.lower().strip() == "plays"), None)
            col_ini    = next((c for c in df.columns if c.lower().strip() == "inicial"), None)
            col_fin    = next((c for c in df.columns if c.lower().strip() == "final"), None)

            if not col_titulo or not col_plays:
                continue

            mask = df[col_titulo].astype(str).str.upper().str.contains(termino_upper, na=False, regex=False)
            filas = df[mask]

            if filas.empty:
                print(f"  ⚠️  {f}: programa no encontrado")
                continue

            row = filas.iloc[0]

            # Usar el datetime completo de la columna Inicial/Final como fuente de verdad.
            # Programas nocturnos tienen fecha de emisión real (ej. 2026-06-12 01:10:00)
            # aunque estén en el archivo del día anterior (2026-06-11).
            ini_raw = str(row[col_ini]) if col_ini and pd.notna(row[col_ini]) else None
            fin_raw = str(row[col_fin]) if col_fin and pd.notna(row[col_fin]) else None

            if ini_raw:
                # Extraer fecha real y hora de la columna Inicial
                ini_dt = datetime.strptime(ini_raw[:19], "%Y-%m-%d %H:%M:%S")
                fecha_emision = ini_dt.strftime("%Y-%m-%d")
                ini_hora = ini_dt.strftime("%H:%M:%S")
            else:
                fecha_emision = str(f)
                ini_hora = None

            fin_hora = str(row[col_fin])[11:19] if fin_raw else None

            resultados.append({
                "fecha":         fecha_emision,   # fecha real de emisión (para Mango)
                "fecha_archivo": str(f),           # fecha del archivo C13
                "dia":           DIAS_ES[datetime.strptime(fecha_emision, "%Y-%m-%d").weekday()],
                "ini":           ini_hora,
                "fin":           fin_hora,
                "plays_s3":      int(float(row[col_plays])) if pd.notna(row[col_plays]) else 0,
            })
        except Exception as e:
            print(f"  ❌ Error leyendo {f}: {e}")

    return resultados

# ─── Paso 2: consultar Mango ──────────────────────────────────────────────────

def consultar_mango(slug, date_start, date_end, reintentos=3):
    payload = {
        "token":      TOKEN,
        "slug":       slug,
        "type":       "overview",
        "format":     "live",
        "date_start": date_start,
        "date_end":   date_end,
    }
    for intento in range(reintentos):
        try:
            r = requests.post(API_URL, json=payload, timeout=60)
            d = r.json()
            if "plays" in d:
                return d
            return None
        except Exception as e:
            if intento < reintentos - 1:
                print(f" reintentando...", end="", flush=True)
                time.sleep(3)
            else:
                print(f" ❌ {e}")
                return None

# ─── Paso 3: exportar Excel ───────────────────────────────────────────────────

def exportar_excel(programa, slug, registros):
    AZUL_OSC = "1F3864"; AZUL_MED = "2E5FA3"
    VERDE_BG = "E2EFDA"; VERDE    = "375623"
    AZUL_BG  = "DDEEFF"; AZUL_FT  = "1F3864"
    BLANCO   = "FFFFFF"; GRIS     = "F5F5F5"

    usar_mango = slug not in ("C13", None)
    fecha_ini  = registros[0]["fecha"]
    fecha_fin  = registros[-1]["fecha"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"C13 vs {slug}" if usar_mango else "Solo C13"

    # Título
    titulo = f"{programa.upper()} — C13 vs {slug}  ({fecha_ini} al {fecha_fin})" if usar_mango \
             else f"{programa.upper()} — Solo C13  ({fecha_ini} al {fecha_fin})"

    # Encabezados: Fecha Emisión | Archivo | Día | Inicio | Fin | Duración | Plays C13 [| Plays {slug}]
    headers = ["Fecha Emisión", "Archivo", "Día", "Inicio", "Fin", "Duración", "Plays C13"]
    if usar_mango:
        headers += [f"Plays {slug}"]
    ncols = chr(64 + len(headers))
    ws.merge_cells(f"A1:{ncols}1")
    ws["A1"] = titulo
    ws["A1"].font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=AZUL_OSC)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=AZUL_MED)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 24

    COL_PLAYS_C13  = 7   # columna G
    COL_PLAYS_SLUG = 8   # columna H

    # Datos
    for ri, reg in enumerate(registros):
        fila = 3 + ri
        bg = BLANCO if ri % 2 == 0 else GRIS
        vals = [reg["fecha"], reg["fecha_archivo"], reg["dia"], reg["ini"], reg["fin"],
                reg["duracion"], reg["plays_s3"]]
        if usar_mango:
            vals += [reg["plays_mango"]]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=fila, column=ci, value=val)
            c.font = Font(name="Calibri", size=10)
            c.alignment = Alignment(horizontal="center" if ci <= 6 else "right")
            if ci == COL_PLAYS_C13:
                c.fill = PatternFill("solid", fgColor=VERDE_BG)
                c.font = Font(name="Calibri", size=10, bold=True, color=VERDE)
                c.number_format = "#,##0"
            elif usar_mango and ci == COL_PLAYS_SLUG:
                c.fill = PatternFill("solid", fgColor=AZUL_BG if val else bg)
                c.font = Font(name="Calibri", size=10, bold=True, color=AZUL_FT if val else "000000")
                c.number_format = "#,##0"
            else:
                c.fill = PatternFill("solid", fgColor=bg)

    # Fila total
    fila_tot = 3 + len(registros)
    ws.merge_cells(f"A{fila_tot}:F{fila_tot}")
    c = ws.cell(row=fila_tot, column=1, value="TOTAL")
    c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=AZUL_OSC)
    c.alignment = Alignment(horizontal="center")
    campos_total = [(COL_PLAYS_C13, "plays_s3")]
    if usar_mango:
        campos_total += [(COL_PLAYS_SLUG, "plays_mango")]
    for ci, campo in campos_total:
        total = sum(r[campo] for r in registros if r[campo])
        c = ws.cell(row=fila_tot, column=ci, value=total)
        c.number_format = "#,##0"
        c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=AZUL_OSC)
        c.alignment = Alignment(horizontal="right")

    # Anchos
    anchos = [13, 13, 6, 8, 8, 10, 13] + ([14] if usar_mango else [])
    for col, ancho in zip("ABCDEFGH", anchos):
        ws.column_dimensions[col].width = ancho

    slug_safe = (slug or "").replace("/", "_")
    prog_safe = programa.replace(" ", "_").replace("/", "_")[:30]
    nombre = f"C13_vs_{slug_safe}_{prog_safe}_{fecha_ini}_{fecha_fin}.xlsx" if usar_mango \
             else f"C13_{prog_safe}_{fecha_ini}_{fecha_fin}.xlsx"
    salida = os.path.join(CARPETA_SALIDA, nombre)
    wb.save(salida)
    return salida

# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    print("\n╔══════════════════════════════════════════╗")
    print("║   C13 vs Mango — Programa por franja    ║")
    print("╚══════════════════════════════════════════╝\n")

    # 1. Rango de fechas
    desde = pedir_fecha("Fecha DESDE")
    hasta = pedir_fecha("Fecha HASTA ")

    # 2. Buscar programa en archivos C13
    print("\n🔍 Buscando programas disponibles en C13...")
    programas = listar_programas_disponibles(desde, hasta)
    if not programas:
        print("❌ No se encontraron archivos de programación en ese rango.")
        return

    termino = pedir("\nNombre o parte del programa a buscar").strip()

    coincidencias = [p for p in programas if termino.upper() in p.upper()]
    if not coincidencias:
        print(f"  ⚠️  Sin coincidencias para '{termino}'. Usando el texto tal cual.")
        termino_busqueda = termino
    elif len(coincidencias) == 1:
        termino_busqueda = coincidencias[0]
        print(f"  ✅ Programa encontrado: {termino_busqueda}")
    else:
        print("\n  Coincidencias encontradas:")
        for i, p in enumerate(coincidencias, 1):
            print(f"    {i}. {p}")
        op = pedir("  Selecciona número").strip()
        try:
            termino_busqueda = coincidencias[int(op) - 1]
        except Exception:
            termino_busqueda = coincidencias[0]
        print(f"  ✅ Seleccionado: {termino_busqueda}")

    # 3. Extraer datos C13
    print(f"\n📂 Extrayendo datos C13...")
    datos_s3 = buscar_programa_en_s3(desde, hasta, termino_busqueda)

    if not datos_s3:
        print("❌ Sin datos C13. Verifica que los archivos estén descargados.")
        return

    dias_sin_hora = [d for d in datos_s3 if not d["ini"] or not d["fin"]]
    if dias_sin_hora:
        print(f"  ⚠️  {len(dias_sin_hora)} días sin columnas Inicial/Final.")

    print(f"  ✅ {len(datos_s3)} días con datos C13")

    # 4. Selección de señal Mango (opcional)
    print("\nSeñales Mango disponibles:")
    for k, v in SENALES_CONOCIDAS.items():
        print(f"  {k}. {v}")
    print("  0. Ingresar slug manualmente")
    print("  N. Solo C13, sin consultar Mango")
    opcion = pedir("Selecciona una opción").strip().upper()

    usar_mango = opcion != "N"
    if usar_mango:
        if opcion == "0" or opcion not in SENALES_CONOCIDAS:
            slug = pedir("Slug de la señal").upper()
        else:
            slug = SENALES_CONOCIDAS[opcion]
        print(f"  ✅ Señal: {slug}")
    else:
        slug = None
        print("  ✅ Solo C13 — sin consulta a Mango")

    # 5. Procesar datos
    print(f"\n{'🌐 Consultando Mango (' + slug + ')...' if usar_mango else '📂 Consolidando datos C13...'}\n")
    registros = []

    for item in datos_s3:
        ini = item["ini"]
        fin = item["fin"]

        if not ini or not fin:
            print(f"  {item['fecha']} ({item['dia']}): sin horario, omitido")
            continue

        t_ini = datetime.strptime(ini, "%H:%M:%S")
        t_fin = datetime.strptime(fin, "%H:%M:%S")
        segundos = (t_fin - t_ini).total_seconds()
        if segundos < 0:
            segundos += 86400
        mins = int(segundos // 60)

        if usar_mango:
            print(f"  {item['fecha']} ({item['dia']}) {ini[:5]}–{fin[:5]}...", end=" ", flush=True)
            datos_mango = consultar_mango(slug, f"{item['fecha']} {ini}", f"{item['fecha']} {fin}")
            plays_m = int(datos_mango.get("plays", 0)) if datos_mango else 0
            print(f"C13: {item['plays_s3']:,}  |  {slug}: {plays_m:,}")
        else:
            print(f"  {item['fecha']} ({item['dia']}) {ini[:5]}–{fin[:5]}  C13: {item['plays_s3']:,}")
            plays_m = None

        registros.append({
            "fecha":         item["fecha"],
            "fecha_archivo": item["fecha_archivo"],
            "dia":           item["dia"],
            "ini":           ini[:5],
            "fin":           fin[:5],
            "duracion":      f"{mins} min",
            "plays_s3":      item["plays_s3"],
            "plays_mango":   plays_m,
        })

    if not registros:
        print("\n❌ Sin registros para exportar.")
        return

    # 6. Resumen
    print(f"\n{'─'*50}")
    print(f"  Total plays C13:  {sum(r['plays_s3'] for r in registros):>10,}")
    if usar_mango:
        print(f"  Total plays {slug}: {sum(r['plays_mango'] for r in registros if r['plays_mango']):>10,}")
    print(f"{'─'*50}")

    # 7. Exportar
    print("\n📝 Generando Excel...")
    salida = exportar_excel(termino_busqueda, slug or "C13", registros)
    print(f"  ✅ {os.path.basename(salida)}")
    print(f"  📍 {CARPETA_SALIDA}\n")

if __name__ == "__main__":
    main()
