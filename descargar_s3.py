"""
descargar_s3.py — Descarga archivos de programación desde S3
=============================================================
Uso:
    python descargar_s3.py                          # te pregunta las fechas
    python descargar_s3.py 2026-04-01 2026-04-30   # rango por argumento

Requiere:
    pip install boto3 openpyxl requests
    AWS configurado: aws configure  (o variables AWS_ACCESS_KEY_ID / AWS_SECRET_ACCESS_KEY)

Si un archivo descargado tiene el Total de Plays en 0, se dispara un reproceso en
Mango (mango.digitalproserver.com/test_canal13.php?day=DD-MM-YYYY), se espera y se
vuelve a descargar. Si sigue en 0, se reporta al final y el archivo no se conserva.
"""

import sys
import re
import time
import boto3
import requests
from datetime import date, timedelta
from pathlib import Path
from openpyxl import load_workbook
from botocore.exceptions import ClientError, NoCredentialsError

try:
    from config import AWS_BUCKET as CONFIG_BUCKET, AWS_REGION as CONFIG_REGION
except Exception:
    CONFIG_BUCKET = None
    CONFIG_REGION = None

# ─── Configuración ────────────────────────────────────────────────────────────
BUCKET = CONFIG_BUCKET or "datos-dps"
REGION = CONFIG_REGION or "us-east-2"
CARPETA_DESTINO = Path(r"C:\procesos\ReporteCanal13\datos")

# Reproceso en Mango para días con el archivo en 0 (total de Plays vacío)
MANGO_REPROCESO_URL = "https://mango.digitalproserver.com/test_canal13.php"
MANGO_ESPERA_SEGUNDOS = 20

# Patrón de carpetas dentro del bucket: YYYY-MM-DD/
PATRON_CARPETA = re.compile(r"^(\d{4}-\d{2}-\d{2})/$")
# Patrón del archivo dentro de cada carpeta: YYYY-MM-DD_programacion.*
PATRON_ARCHIVO = re.compile(r"^\d{4}-\d{2}-\d{2}_programacion\.", re.IGNORECASE)


def pedir_fecha(mensaje: str) -> date:
    while True:
        texto = input(mensaje).strip()
        try:
            return date.fromisoformat(texto)
        except ValueError:
            print("  Formato incorrecto. Usa YYYY-MM-DD (ej. 2026-04-01)")


def fechas_en_rango(desde: date, hasta: date):
    d = desde
    while d <= hasta:
        yield d
        d += timedelta(days=1)


def total_plays(ruta: Path):
    """Lee la fila 'Total' del archivo y retorna el valor de Plays (None si no se pudo leer)."""
    try:
        wb = load_workbook(ruta, read_only=True, data_only=True)
        ws = wb.active
        for fila in ws.iter_rows(min_row=2, values_only=True):
            titulo = fila[2] if len(fila) > 2 else None
            plays = fila[3] if len(fila) > 3 else None
            if titulo and str(titulo).strip().lower() == "total":
                wb.close()
                return float(plays) if plays is not None else 0.0
        wb.close()
    except Exception:
        return None
    return None


def reprocesar_mango(fecha: date) -> bool:
    """Dispara el reproceso en Mango para el día indicado."""
    url = f"{MANGO_REPROCESO_URL}?day={fecha.strftime('%d-%m-%Y')}"
    try:
        r = requests.get(url, timeout=60)
        return r.status_code == 200
    except requests.RequestException as e:
        print(f" ERROR al reprocesar: {e}")
        return False


def descargar_rango(desde: date, hasta: date):
    CARPETA_DESTINO.mkdir(parents=True, exist_ok=True)

    try:
        s3 = boto3.client("s3", region_name=REGION)
    except NoCredentialsError:
        print("ERROR: No se encontraron credenciales AWS.")
        print("Ejecuta 'aws configure' o define AWS_ACCESS_KEY_ID y AWS_SECRET_ACCESS_KEY.")
        sys.exit(1)

    # Construir conjunto de fechas solicitadas como strings "YYYY-MM-DD"
    fechas = {str(f) for f in fechas_en_rango(desde, hasta)}

    print(f"\nBuscando archivos del {desde} al {hasta} en s3://{BUCKET}/...")

    # Listar todas las carpetas del bucket (prefijos de nivel 1)
    paginator = s3.get_paginator("list_objects_v2")
    pages = paginator.paginate(Bucket=BUCKET, Delimiter="/")

    carpetas_encontradas = []
    for page in pages:
        for prefix_obj in page.get("CommonPrefixes", []):
            prefix = prefix_obj["Prefix"]          # ej. "2026-04-01/"
            m = PATRON_CARPETA.match(prefix)
            if m and m.group(1) in fechas:
                carpetas_encontradas.append((m.group(1), prefix))

    if not carpetas_encontradas:
        print("No se encontraron carpetas para el rango indicado.")
        return

    carpetas_encontradas.sort()
    print(f"Carpetas encontradas: {len(carpetas_encontradas)}\n")

    descargados = 0
    omitidos    = 0
    errores     = 0
    vacios      = []

    for fecha_str, prefix in carpetas_encontradas:
        # Listar archivos dentro de la carpeta
        resp = s3.list_objects_v2(Bucket=BUCKET, Prefix=prefix)
        objetos = resp.get("Contents", [])

        archivos_prog = [
            o["Key"] for o in objetos
            if PATRON_ARCHIVO.match(Path(o["Key"]).name)
        ]

        if not archivos_prog:
            print(f"  [{fecha_str}] Sin archivo _programacion — omitido")
            omitidos += 1
            continue

        for key in archivos_prog:
            nombre_archivo = Path(key).name
            destino = CARPETA_DESTINO / nombre_archivo

            if destino.exists():
                print(f"  [{fecha_str}] Ya existe: {nombre_archivo} — omitido")
                omitidos += 1
                continue

            try:
                print(f"  [{fecha_str}] Descargando {nombre_archivo}...", end=" ", flush=True)
                s3.download_file(BUCKET, key, str(destino))
                print("OK")
            except ClientError as e:
                print(f"ERROR: {e}")
                errores += 1
                continue

            total = total_plays(destino)

            if total == 0:
                print(f"  [{fecha_str}] Total en 0 — reprocesando en Mango...", end=" ", flush=True)
                fecha_dt = date.fromisoformat(fecha_str)
                if reprocesar_mango(fecha_dt):
                    print(f"OK, esperando {MANGO_ESPERA_SEGUNDOS}s...")
                    time.sleep(MANGO_ESPERA_SEGUNDOS)
                    try:
                        s3.download_file(BUCKET, key, str(destino))
                        total = total_plays(destino)
                    except ClientError as e:
                        print(f"  [{fecha_str}] ERROR al re-descargar: {e}")
                        errores += 1
                        continue
                else:
                    print("falló la solicitud de reproceso")

            if total == 0:
                print(f"  [{fecha_str}] Sigue en 0 tras reprocesar — no se guarda archivo vacío")
                destino.unlink(missing_ok=True)
                vacios.append(fecha_str)
                continue

            descargados += 1

    print(f"\nResumen: {descargados} descargados, {omitidos} omitidos, {errores} errores, {len(vacios)} vacíos.")
    print(f"Destino: {CARPETA_DESTINO}")
    if vacios:
        print("\n⚠️  Días con Total en 0 incluso tras reprocesar (no se descargaron):")
        for f in vacios:
            print(f"  - {f}")


def main():
    args = sys.argv[1:]

    if len(args) == 2:
        try:
            desde = date.fromisoformat(args[0])
            hasta = date.fromisoformat(args[1])
        except ValueError:
            print("Fechas inválidas. Usa formato YYYY-MM-DD.")
            sys.exit(1)
    elif len(args) == 0:
        print("=== Descarga de archivos S3 — Canal 13 ===\n")
        desde = pedir_fecha("Fecha DESDE (YYYY-MM-DD): ")
        hasta = pedir_fecha("Fecha HASTA  (YYYY-MM-DD): ")
    else:
        print("Uso: python descargar_s3.py [FECHA_DESDE FECHA_HASTA]")
        sys.exit(1)

    if desde > hasta:
        print("La fecha DESDE debe ser anterior o igual a HASTA.")
        sys.exit(1)

    descargar_rango(desde, hasta)


if __name__ == "__main__":
    main()
